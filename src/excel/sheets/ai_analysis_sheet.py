from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.models import PortfolioSummary

class AIAnalysisSheet:
    def __init__(self):
        self.sheet_name = "AI 포트폴리오 진단"
        self.header_font = Font(name='Malgun Gothic', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.content_font = Font(name='Malgun Gothic', size=11)
        self.bold_font = Font(name='Malgun Gothic', size=11, bold=True)
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

    def write(self, workbook, data: PortfolioSummary):
        """
        AI 분석 결과를 엑셀 시트에 작성
        """
        if not data.analysis or "analysis" not in data.analysis:
            return

        ws = workbook.create_sheet(self.sheet_name)
        analysis_text = data.analysis["analysis"]

        # 제목
        ws.merge_cells('B2:H2')
        cell = ws['B2']
        cell.value = "AI 포트폴리오 심층 진단 리포트"
        cell.font = Font(name='Malgun Gothic', size=16, bold=True)
        cell.alignment = self.center_alignment

        # 분석 일시
        ws.merge_cells('B3:H3')
        cell = ws['B3']
        from datetime import datetime
        cell.value = f"진단 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(name='Malgun Gothic', size=10, color='666666')
        cell.alignment = Alignment(horizontal='right')

        # 본문 파싱 및 작성
        current_row = 5
        lines = analysis_text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue

            # 헤더 처리 (###, ##, #)
            if line.startswith('#'):
                current_row += 1
                level = line.count('#')
                text = line.replace('#', '').strip()
                
                ws.merge_cells(f'B{current_row}:H{current_row}')
                cell = ws[f'B{current_row}']
                cell.value = text
                
                if level == 1:
                    cell.font = Font(name='Malgun Gothic', size=14, bold=True, color='2F5597')
                    cell.border = Border(bottom=Side(style='medium', color='2F5597'))
                elif level == 2:
                    cell.font = Font(name='Malgun Gothic', size=13, bold=True, color='2F5597')
                    cell.border = Border(bottom=Side(style='thin', color='2F5597'))
                else:
                    cell.font = Font(name='Malgun Gothic', size=12, bold=True, color='44546A')
                
                cell.alignment = Alignment(vertical='center')
                ws.row_dimensions[current_row].height = 30
                current_row += 1

            # 리스트 아이템 처리 (- , *)
            elif line.startswith('- ') or line.startswith('* '):
                text = line[2:].strip()
                ws.merge_cells(f'C{current_row}:H{current_row}')
                
                # 불릿 포인트
                bullet_cell = ws[f'B{current_row}']
                bullet_cell.value = "•"
                bullet_cell.alignment = Alignment(horizontal='right', vertical='top')
                bullet_cell.font = self.content_font
                
                # 내용 (볼드 처리 포함)
                content_cell = ws[f'C{current_row}']
                self._write_formatted_text(content_cell, text)
                content_cell.alignment = self.left_alignment
                
                # 높이 자동 조절 (대략적인 계산)
                line_count = (len(text) // 60) + 1
                ws.row_dimensions[current_row].height = line_count * 20
                current_row += 1

            # 숫자 리스트 처리 (1. )
            elif line[0].isdigit() and line[1:3] == '. ':
                text = line[3:].strip()
                number = line.split('.')[0]
                
                ws.merge_cells(f'C{current_row}:H{current_row}')
                
                # 번호
                num_cell = ws[f'B{current_row}']
                num_cell.value = f"{number}."
                num_cell.alignment = Alignment(horizontal='right', vertical='top')
                num_cell.font = Font(name='Malgun Gothic', size=11, bold=True, color='4472C4')
                
                # 내용
                content_cell = ws[f'C{current_row}']
                self._write_formatted_text(content_cell, text)
                content_cell.alignment = self.left_alignment
                
                line_count = (len(text) // 60) + 1
                ws.row_dimensions[current_row].height = line_count * 20
                current_row += 1

            # 일반 텍스트
            else:
                ws.merge_cells(f'B{current_row}:H{current_row}')
                cell = ws[f'B{current_row}']
                self._write_formatted_text(cell, line)
                cell.alignment = self.left_alignment
                
                line_count = (len(line) // 70) + 1
                ws.row_dimensions[current_row].height = line_count * 20
                current_row += 1

        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 5  # 불릿/번호 영역
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

    def _write_formatted_text(self, cell, text):
        """
        **bold** 텍스트 처리 (단순화를 위해 전체 텍스트에 적용하거나 제거)
        openpyxl은 셀 내 부분 스타일링이 복잡하므로, **를 제거하고 텍스트만 넣되
        전체가 **로 감싸져 있으면 볼드 처리
        """
        if text.startswith('**') and text.endswith('**'):
            cell.value = text[2:-2]
            cell.font = self.bold_font
        else:
            # 중간에 있는 ** 제거 (가독성을 위해)
            cell.value = text.replace('**', '')
            cell.font = self.content_font

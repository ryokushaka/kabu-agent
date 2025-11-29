from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference, Series
from src.models import PortfolioSummary

def write_analysis_sheet(wb: Workbook, data: PortfolioSummary):
    """포트폴리오 분석 시트 생성 (차트 포함)"""
    ws = wb.create_sheet("포트폴리오 분석") # 순서대로 배치
    
    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid') # 네이비 블루
    row_font = Font(name='맑은 고딕', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. 자산 배분 (Asset Allocation)
    ws['B2'] = "자산 배분 (Asset Allocation)"
    ws['B2'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    headers = ["자산군", "평가금액($)", "비중(%)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=2+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    total_assets = data.total_value_usd + data.cash_balance
    cash_weight = (data.cash_balance / total_assets) if total_assets > 0 else 0
    stock_weight = (data.total_value_usd / total_assets) if total_assets > 0 else 0
    
    assets_data = [
        ["주식 (Stocks)", data.total_value_usd, stock_weight],
        ["현금 (Cash)", data.cash_balance, cash_weight]
    ]
    
    for r_idx, row in enumerate(assets_data, 5):
        for c_idx, val in enumerate(row, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = row_font
            cell.border = thin_border
            if c_idx == 3: # 금액
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif c_idx == 4: # 비중
                cell.number_format = '0.00%'
                cell.alignment = right_align
            else:
                cell.alignment = center_align

    # 자산 배분 차트
    pie = PieChart()
    pie.title = "자산 배분"
    labels = Reference(ws, min_col=2, min_row=5, max_row=6)
    data_ref = Reference(ws, min_col=3, min_row=5, max_row=6)
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(labels)
    ws.add_chart(pie, "E4")

    # 2. 섹터 비중 (Sector Breakdown)
    ws['B15'] = "섹터 비중 (Sector Breakdown)"
    ws['B15'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    headers = ["섹터", "평가금액($)", "비중(%)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=17, column=2+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    sector_data = []
    if data.sector_allocation:
        for sector, value in data.sector_allocation.items():
            weight = (value / data.total_value_usd) if data.total_value_usd > 0 else 0
            sector_data.append([sector, value, weight])
    else:
        sector_data.append(["Unknown", data.total_value_usd, 1.0])
        
    # Sort by value desc
    sector_data.sort(key=lambda x: x[1], reverse=True)
    
    for r_idx, row in enumerate(sector_data, 18):
        for c_idx, val in enumerate(row, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = row_font
            cell.border = thin_border
            if c_idx == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif c_idx == 4:
                cell.number_format = '0.00%'
                cell.alignment = right_align
            else:
                cell.alignment = center_align
                
    # 섹터 차트
    pie2 = PieChart()
    pie2.title = "섹터 비중"
    labels = Reference(ws, min_col=2, min_row=18, max_row=17+len(sector_data))
    data_ref = Reference(ws, min_col=3, min_row=18, max_row=17+len(sector_data))
    pie2.add_data(data_ref, titles_from_data=False)
    pie2.set_categories(labels)
    ws.add_chart(pie2, "E17")
    
    # 3. 상위 보유 종목 (Top 5 Holdings)
    ws['B35'] = "상위 5개 보유 종목 (Top 5 Holdings)"
    ws['B35'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    headers = ["종목명", "평가금액($)", "비중(%)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=37, column=2+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Sort holdings by value
    sorted_holdings = sorted(data.holdings, key=lambda x: x.total_value, reverse=True)[:5]
    
    for r_idx, h in enumerate(sorted_holdings, 38):
        weight = (h.total_value / data.total_value_usd) if data.total_value_usd > 0 else 0
        row = [h.name, h.total_value, weight]
        for c_idx, val in enumerate(row, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = row_font
            cell.border = thin_border
            if c_idx == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif c_idx == 4:
                cell.number_format = '0.00%'
                cell.alignment = right_align
            else:
                cell.alignment = center_align

    # 상위 종목 바 차트
    bar = BarChart()
    bar.title = "상위 5개 종목"
    bar.y_axis.title = '평가금액 ($)'
    labels = Reference(ws, min_col=2, min_row=38, max_row=37+len(sorted_holdings))
    data_ref = Reference(ws, min_col=3, min_row=38, max_row=37+len(sorted_holdings))
    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(labels)
    ws.add_chart(bar, "E35")

    # 컬럼 너비 조정
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

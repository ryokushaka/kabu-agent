from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.models import PortfolioSummary
from datetime import datetime

def write_cover_sheet(wb: Workbook, data: PortfolioSummary):
    """표지 시트 생성"""
    if "Cover" in wb.sheetnames:
        ws = wb["Cover"]
    else:
        ws = wb.create_sheet("Cover")
    
    # 스타일 정의
    title_font = Font(name='맑은 고딕', size=24, bold=True, color='203764')
    subtitle_font = Font(name='맑은 고딕', size=14, color='555555')
    info_font = Font(name='맑은 고딕', size=12)
    
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 타이틀
    ws.merge_cells('B5:H7')
    cell = ws['B5']
    cell.value = "Investment Portfolio Report"
    cell.font = title_font
    cell.alignment = center_align
    
    # 서브 타이틀 (날짜)
    ws.merge_cells('B8:H9')
    cell = ws['B8']
    cell.value = datetime.now().strftime("%Y년 %m월 %d일")
    cell.font = subtitle_font
    cell.alignment = center_align
    
    # 고객 정보 (임시)
    ws.merge_cells('B12:H12')
    ws['B12'] = "Client: Admin User"
    ws['B12'].font = info_font
    ws['B12'].alignment = center_align
    
    ws.merge_cells('B13:H13')
    ws['B13'] = f"Total AUM: ${data.total_value_usd + data.cash_balance:,.2f}"
    ws['B13'].font = info_font
    ws['B13'].alignment = center_align
    
    # 면책 조항
    ws.merge_cells('B20:H22')
    disclaimer = "본 리포트는 정보 제공을 목적으로 하며, 투자의 책임은 투자자 본인에게 있습니다.\n(This report is for informational purposes only.)"
    cell = ws['B20']
    cell.value = disclaimer
    cell.font = Font(name='맑은 고딕', size=9, italic=True, color='888888')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 배경색 (선택사항)
    # ws.sheet_view.showGridLines = False

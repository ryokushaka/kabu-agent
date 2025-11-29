from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.models import PortfolioSummary
from datetime import datetime

def write_dashboard_sheet(wb: Workbook, data: PortfolioSummary):
    """대시보드(요약) 시트 생성"""
    ws = wb.create_sheet("포트폴리오 요약")
    
    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    label_font = Font(name='맑은 고딕', size=11, bold=True)
    value_font = Font(name='맑은 고딕', size=11)
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 제목
    ws['B2'] = "해외주식 포트폴리오 요약 보고서"
    ws['B2'].font = Font(name='맑은 고딕', size=20, bold=True, color='1F4E78')
    
    # 생성 일시
    ws['B3'] = f"생성 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['B3'].font = Font(name='맑은 고딕', size=10, color='7F7F7F')

    # 주요 지표 테이블 헤더
    headers = ["항목", "값"]
    ws['B5'] = headers[0]
    ws['C5'] = headers[1]
    
    for col in ['B', 'C']:
        cell = ws[f'{col}5']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 데이터 매핑
    metrics = [
        ("총 자산 (USD)", f"${data.total_value_usd:,.2f}"),
        ("총 자산 (KRW)", f"₩{data.total_value_krw:,.0f}"),
        ("총 평가 손익", f"${data.total_profit_loss:,.2f}"),
        ("총 수익률", f"{data.total_return_rate:.2f}%"),
        ("보유 종목 수", f"{len(data.holdings)} 개")
    ]

    # 데이터 입력
    start_row = 6
    for i, (label, value) in enumerate(metrics):
        row = start_row + i
        ws[f'B{row}'] = label
        ws[f'C{row}'] = value
        
        ws[f'B{row}'].font = label_font
        ws[f'B{row}'].alignment = center_align
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'].font = value_font
        ws[f'C{row}'].alignment = right_align
        ws[f'C{row}'].border = thin_border
        
        # 수익률/손익 색상 처리
        if "수익률" in label or "손익" in label:
            if "-" in value:
                ws[f'C{row}'].font = Font(name='맑은 고딕', size=11, color='FF0000') # 파란색 (손실) - 한국 기준 파란색이 하락이나 미국 기준 빨간색이 하락. 
                # 컨설턴트 요청사항: 가독성 좋은 양식. 통상적으로 엑셀에서는 빨간색이 음수 표현에 많이 쓰임.
                # 여기서는 빨간색을 음수로 사용.
            else:
                ws[f'C{row}'].font = Font(name='맑은 고딕', size=11, color='0000FF') # 빨간색 (상승) - 한국 기준 빨간색이 상승.
                # 엑셀 기본 서식 따름: 음수 빨강, 양수 검정/파랑. 
                # 한국 주식 시장 관행: 상승 빨강, 하락 파랑.
                # 미국 주식 시장 관행: 상승 초록, 하락 빨강.
                # 한국 사용자 대상이므로 상승 빨강(FF0000), 하락 파랑(0000FF)으로 가겠음.
                
                # 수정: 위 코드에서 FF0000은 빨강, 0000FF는 파랑.
                # 손실(음수) -> 파랑 (0000FF)
                # 이익(양수) -> 빨강 (FF0000)
                
                if "-" in value:
                     ws[f'C{row}'].font = Font(name='맑은 고딕', size=11, color='0000FF') # 파랑
                else:
                     ws[f'C{row}'].font = Font(name='맑은 고딕', size=11, color='FF0000') # 빨강

    # 컬럼 너비 조정
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    
    # 눈금선 제거
    ws.sheet_view.showGridLines = False
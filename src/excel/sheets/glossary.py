from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def write_glossary_sheet(wb: Workbook):
    """용어 설명(Glossary) 시트 생성"""
    ws = wb.create_sheet("용어 설명 (Glossary)")
    
    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='555555', end_color='555555', fill_type='solid')
    term_font = Font(name='맑은 고딕', size=11, bold=True)
    desc_font = Font(name='맑은 고딕', size=10)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 제목
    ws['B2'] = "금융 용어 설명 (Financial Glossary)"
    ws['B2'].font = Font(name='맑은 고딕', size=16, bold=True)
    
    # 테이블 헤더
    ws['B4'] = "용어 (Term)"
    ws['C4'] = "설명 (Description)"
    
    for col in ['B', 'C']:
        cell = ws[f'{col}4']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 용어 데이터
    terms = [
        ("변동성 (Volatility)", "자산 가격의 변동 폭을 나타내는 지표로, 수치가 높을수록 위험도가 높음을 의미합니다. (연간화된 표준편차)"),
        ("베타 (Beta)", "시장 전체(S&P 500) 대비 개별 포트폴리오의 민감도를 나타냅니다. 1보다 크면 시장보다 변동성이 크고, 1보다 작으면 적음을 의미합니다."),
        ("샤프 지수 (Sharpe Ratio)", "위험 1단위당 초과 수익률을 나타냅니다. 수치가 높을수록 위험 대비 성과가 우수함을 의미합니다."),
        ("최대 낙폭 (Max Drawdown)", "특정 기간 동안 고점 대비 가장 많이 하락한 비율을 의미합니다. 손실 위험의 최대치를 가늠할 수 있습니다."),
        ("S&P 500", "미국 주식시장을 대표하는 500개 대형기업의 주가 지수입니다. 벤치마크(비교 기준)로 사용됩니다."),
        ("섹터 (Sector)", "주식 시장을 산업별로 분류한 그룹입니다. (예: 기술, 헬스케어, 금융 등)"),
        ("평가 손익 (Unrealized P/L)", "현재 보유 중인 자산의 가치 변동에 따른 잠재적 이익이나 손실입니다."),
        ("수익률 (Return Rate)", "투자 원금 대비 수익(또는 손실)의 비율을 백분율로 나타낸 것입니다.")
    ]
    
    # 데이터 입력
    start_row = 5
    for i, (term, desc) in enumerate(terms):
        row = start_row + i
        ws[f'B{row}'] = term
        ws[f'C{row}'] = desc
        
        ws[f'B{row}'].font = term_font
        ws[f'B{row}'].alignment = center_align
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'].font = desc_font
        ws[f'C{row}'].alignment = left_align
        ws[f'C{row}'].border = thin_border
        
    # 컬럼 너비 조정
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    
    # 눈금선 제거
    ws.sheet_view.showGridLines = False

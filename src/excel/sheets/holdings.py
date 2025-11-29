from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.models import PortfolioSummary

def write_holdings_sheet(wb: Workbook, data: PortfolioSummary):
    """보유 종목 상세 시트 생성"""
    ws = wb.create_sheet("보유 종목 상세")
    
    # 스타일 정의
    title_font = Font(name='맑은 고딕', size=16, bold=True, color='203764')
    info_font = Font(name='맑은 고딕', size=10, color='555555')
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    row_font = Font(name='맑은 고딕', size=10)
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. 제목 및 환율 정보
    ws['B2'] = "보유 종목 상세 분석"
    ws['B2'].font = title_font
    
    from datetime import datetime
    today_str = datetime.now().strftime("%Y-%m-%d")
    ws['B3'] = f"기준 환율: 1 USD = {data.exchange_rate:,.2f} KRW  |  기준일: {today_str}"
    ws['B3'].font = info_font

    # 2. 헤더
    headers = ["종목코드", "종목명", "섹터", "보유수량", "비중(%)", "평균단가($)", "현재가($)", "52주 최고($)", "52주 최저($)", "평가금액($)", "평가손익($)", "수익률(%)", "매수일자", "보유일수"]
    
    start_row = 5
    for col_idx, header in enumerate(headers, 2): # B열부터 시작
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 3. 데이터 입력
    for row_idx, holding in enumerate(data.holdings, start_row + 1):
        row_data = [
            holding.symbol,
            holding.name,
            holding.sector,
            holding.quantity,
            holding.weight / 100,
            holding.average_price,
            holding.current_price,
            holding.fifty_two_week_high,
            holding.fifty_two_week_low,
            holding.total_value,
            holding.total_value - (holding.average_price * holding.quantity),
            holding.return_rate / 100,
            "-", # 매수일자 (정보 없음)
            "-"  # 보유일수 (정보 없음)
        ]
        
        for i, value in enumerate(row_data):
            col_idx = 2 + i
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.border = thin_border
            
            # 정렬 및 포맷팅
            if i in [0, 1, 2, 12, 13]: # 코드, 이름, 섹터, 날짜, 일수
                cell.alignment = center_align
            else: # 숫자 데이터
                cell.alignment = right_align
                if i in [5, 6, 7, 8, 9, 10]: # 금액 ($)
                    cell.number_format = '#,##0.00'
                elif i == 3: # 수량
                    cell.number_format = '#,##0'
                elif i in [4, 11]: # 비중, 수익률 (%)
                    cell.number_format = '0.00%'
                    
            # 수익률/손익 색상
            if i in [10, 11]: # 평가손익, 수익률
                if isinstance(value, (int, float)):
                    if value > 0:
                        cell.font = Font(name='맑은 고딕', size=10, color='FF0000') # 빨강
                    elif value < 0:
                        cell.font = Font(name='맑은 고딕', size=10, color='0000FF') # 파랑

    # 컬럼 너비 자동 조정
    widths = [10, 25, 15, 10, 10, 12, 12, 12, 12, 15, 15, 12, 12, 10]
    for i, width in enumerate(widths):
        ws.column_dimensions[chr(66+i)].width = width # B열(66)부터 시작
        
    ws.sheet_view.showGridLines = False
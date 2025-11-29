from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.models import PortfolioSummary

def write_returns_sheet(wb: Workbook, data: PortfolioSummary):
    """수익률 분석 시트 생성"""
    ws = wb.create_sheet("수익률 분석")
    
    ws['B2'] = "포트폴리오 수익률 분석"
    ws['B2'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    # 간단한 수익률 요약 테이블
    headers = ["구분", "수익률(%)", "평가손익($)"]
    
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_idx, header in enumerate(headers, 2):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    # 현재는 전체 수익률만 표시 (추후 기간별 수익률 데이터가 있으면 추가)
    row_data = ["전체 기간", data.total_return_rate / 100, data.total_profit_loss]
    
    ws.cell(row=5, column=2, value=row_data[0]).border = thin_border
    
    rate_cell = ws.cell(row=5, column=3, value=row_data[1])
    rate_cell.number_format = '0.00%'
    rate_cell.border = thin_border
    
    val_cell = ws.cell(row=5, column=4, value=row_data[2])
    val_cell.number_format = '#,##0.00'
    val_cell.border = thin_border
    
    # 색상 적용
    if data.total_profit_loss > 0:
        color = 'FF0000'
    elif data.total_profit_loss < 0:
        color = '0000FF'
    else:
        color = '000000'
        
    rate_cell.font = Font(color=color)
    val_cell.font = Font(color=color)

    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    ws.sheet_view.showGridLines = False

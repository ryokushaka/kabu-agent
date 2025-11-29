from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from src.models import PortfolioSummary

def write_risk_sheet(wb: Workbook, data: PortfolioSummary):
    """리스크 분석 시트 생성"""
    ws = wb.create_sheet("리스크 분석") # Cover 다음
    
    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
    row_font = Font(name='맑은 고딕', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. 리스크 지표 (Risk Metrics)
    ws['B2'] = "리스크 지표 (Risk Metrics)"
    ws['B2'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    metrics_headers = ["지표 (Metric)", "값 (Value)", "설명 (Description)"]
    for i, h in enumerate(metrics_headers):
        cell = ws.cell(row=4, column=2+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    metrics = data.analysis.get("metrics", {}) if data.analysis else {}
    
    metrics_data = [
        ["변동성 (Volatility)", metrics.get("volatility", 0), "연간화된 표준편차 (위험도)"],
        ["베타 (Beta)", metrics.get("beta", 0), "시장(S&P500) 민감도"],
        ["샤프 지수 (Sharpe Ratio)", metrics.get("sharpe_ratio", 0), "위험 대비 수익률 (높을수록 좋음)"],
        ["최대 낙폭 (Max Drawdown)", metrics.get("max_drawdown", 0), "고점 대비 최대 하락폭"]
    ]
    
    for r_idx, row in enumerate(metrics_data, 5):
        for c_idx, val in enumerate(row, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = row_font
            cell.border = thin_border
            if c_idx == 3: # 값
                cell.number_format = '0.00'
                if row[0] == "최대 낙폭 (Max Drawdown)":
                     cell.number_format = '0.00%'
                cell.alignment = right_align
            else:
                cell.alignment = center_align
                
    # 2. 벤치마크 비교 (Benchmark Comparison)
    ws['B12'] = "벤치마크 비교 (Benchmark Comparison)"
    ws['B12'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    chart_data = data.analysis.get("chart_data", []) if data.analysis else []
    
    # 데이터 테이블 작성 (숨김 처리하거나 아래쪽에 배치)
    # 차트를 위해 데이터가 시트에 있어야 함
    start_row = 15
    ws.cell(row=start_row, column=2, value="Date")
    ws.cell(row=start_row, column=3, value="Portfolio (%)")
    ws.cell(row=start_row, column=4, value="S&P 500 (%)")
    
    for i, item in enumerate(chart_data):
        ws.cell(row=start_row+1+i, column=2, value=item["date"])
        ws.cell(row=start_row+1+i, column=3, value=item["portfolio"])
        ws.cell(row=start_row+1+i, column=4, value=item["benchmark"])
        
    # 라인 차트 생성
    chart = LineChart()
    chart.title = "Portfolio vs S&P 500 (Cumulative Return)"
    chart.style = 13
    chart.y_axis.title = 'Return (%)'
    chart.x_axis.title = 'Date'
    
    data_ref = Reference(ws, min_col=3, min_row=start_row, max_col=4, max_row=start_row+len(chart_data))
    cats_ref = Reference(ws, min_col=2, min_row=start_row+1, max_row=start_row+len(chart_data))
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    ws.add_chart(chart, "E4") # 차트 위치
    
    # 컬럼 너비 조정
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40

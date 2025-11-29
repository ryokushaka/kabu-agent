from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.models import PortfolioSummary

def write_sector_sheet(wb: Workbook, data: PortfolioSummary):
    """섹터 분석 시트 생성"""
    ws = wb.create_sheet("섹터 분석", 2)
    
    # TODO: 실제 섹터 데이터가 모델에 없으므로, 현재는 임시로 종목별 비중으로 대체하거나 
    # 추후 API에서 섹터 정보를 가져오도록 수정 필요.
    # 여기서는 '종목별 비중'으로 대체하여 구현.
    
    ws['B2'] = "종목별 비중 분석"
    ws['B2'].font = Font(name='맑은 고딕', size=14, bold=True)
    
    headers = ["종목명", "평가금액($)", "비중(%)"]
    
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_idx, header in enumerate(headers, 2): # B열부터 시작
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    total_value = data.total_value_usd
    
    for row_idx, holding in enumerate(data.holdings, 5):
        weight = (holding.total_value / total_value) if total_value > 0 else 0
        
        ws.cell(row=row_idx, column=2, value=holding.name).border = thin_border
        
        val_cell = ws.cell(row=row_idx, column=3, value=holding.total_value)
        val_cell.number_format = '#,##0.00'
        val_cell.border = thin_border
        
        weight_cell = ws.cell(row=row_idx, column=4, value=weight)
        weight_cell.number_format = '0.00%'
        weight_cell.border = thin_border

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    ws.sheet_view.showGridLines = False
